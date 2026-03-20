"""
migrate_excel.py
────────────────
Excelの販売管理シートをAirtable sales_management テーブルに移行する。

使い方:
    python scripts/migrate_excel.py
"""
import json, os, sys, urllib.request, time
from pathlib import Path
from datetime import datetime

import openpyxl

# 環境変数 COIN_EXCEL_PATH で上書き可能（デフォルト: ~/Desktop/コイン販売管理シート.xlsx）
EXCEL_PATH = os.environ.get("COIN_EXCEL_PATH", str(Path.home() / "Desktop" / "コイン販売管理シート.xlsx"))
ENV_PATH = Path(__file__).resolve().parent.parent / ".env"

def load_env():
    env = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        if line.strip() and "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env

ENV = load_env()
API_KEY = ENV["AIRTABLE_API_KEY"]
BASE_ID = ENV["AIRTABLE_BASE_ID"]
TBL_SM = "tbllEIckahvgklrZ7"

# field IDs
F = {
    "no": "fldPnBH0XfzTe51UJ",
    "coin_id": "fldqly5OojXSB4g9d",
    "material": "flds1Y5vCXEr5atDr",
    "cert_no": "fldtDk7ejtZKriUpQ",
    "item_name": "flda2YulhLJkhWVHX",
    "purchase_date": "fldTyGiIF5KoHOMfO",
    "source": "fldNSkWkm8gzND1Qd",
    "currency": "fldKsAPq18vGbtEay",
    "exchange_rate": "fldxk1x9hvYQuIMLM",
    "price_foreign": "flduJjxz0HzeN1XZ3",
    "price_jpy": "fldPGxnzMvFf67XWd",
    "commission_foreign": "fldWnm0oNHIYo9qZu",
    "commission_jpy": "fldcAh7Enc3E1g95Q",
    "tax": "fldWGWR77nZKaxjpa",
    "shipping_purchase": "fld1iFafqyZWKHg0m",
    "other_purchase": "fldTY5UaNu2tUDTR2",
    "total_purchase": "fldWYo28LoFiqu0SM",
    "sale_date": "fldyBw6cSK6zbwiCL",
    "revenue": "fldRlyBCzBkXkl17Z",
    "sale_commission": "fldRhuJoVx0Qk8BJ1",
    "shipping_sale": "fldL5a9BwVxTP4mnA",
    "other_sale": "fldklsqfX596BtO64",
    "total_sale": "fldAKpLYjUyW7oaR0",
    "profit": "fldzIoTOcug8LybWn",
    "title": "fldNuogDP5uEeeRR3",
    "description": "fldXYw84VgT9cAN04",
    "min_price": "fldB1pJlu1p0Z1qrU",
    "min_profit": "flduMdbW6rVbOx8gW",
    "buynow_price": "fldVaT8Q3JO1Ktwez",
    "buynow_profit": "fldRDtPV9KsKrm70g",
    "status": "fldOUwU69T50Z7O90",
    "ebay_url": "fldHq48sF0jwBRh0K",
    "memo": "fld2XHOKSS4Pr1GbE",
    "photo": "flduEaSvudpT2bC9G",
}


def api_post(records):
    url = f"https://api.airtable.com/v0/{BASE_ID}/{TBL_SM}"
    data = json.dumps({"records": records}).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    })
    try:
        resp = urllib.request.urlopen(req)
        return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        print(f"  ERROR {e.code}: {body[:300]}")
        return None


def safe_num(val):
    """数値変換。エラーやNoneは0"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val)
    try:
        return round(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return None


def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    try:
        return round(float(str(val).replace(",", "")), 2)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def parse_status(val):
    """出品ステータスを判定"""
    s = safe_str(val).strip()
    if not s:
        return "未出品"
    if "出品" in s:
        return "出品中"
    if "売却" in s or "済" in s:
        return "売却済"
    return "未出品"


def migrate_sheet(ws, sheet_name, start_row, col_map, is_sold=False):
    """シートからデータを読み込んでAirtableに投入"""
    records = []
    count = 0

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=False):
        no_val = row[col_map["no"]].value
        item_val = row[col_map["item"]].value

        if no_val is None or not isinstance(no_val, (int, float)):
            continue
        if item_val is None:
            continue

        no = int(no_val)
        coin_id = f"EXCEL-{no:03d}" if not is_sold else f"SOLD-{no:03d}"

        fields = {}
        fields[F["no"]] = no
        fields[F["coin_id"]] = coin_id
        fields[F["item_name"]] = safe_str(item_val)[:200]

        # 素材
        mat = safe_str(row[col_map.get("material", 99)].value if col_map.get("material") is not None else "")
        if mat:
            fields[F["material"]] = mat

        # 鑑定番号
        cert = safe_str(row[col_map.get("cert_no", 99)].value if col_map.get("cert_no") is not None else "")
        if cert:
            fields[F["cert_no"]] = cert

        # 仕入データ
        source = safe_str(row[col_map["source"]].value)
        if source:
            fields[F["source"]] = source

        currency = safe_str(row[col_map["currency"]].value)
        if currency:
            fields[F["currency"]] = currency

        rate = safe_float(row[col_map["exchange_rate"]].value)
        if rate is not None:
            fields[F["exchange_rate"]] = rate

        fp = safe_float(row[col_map["price_foreign"]].value)
        if fp is not None:
            fields[F["price_foreign"]] = fp

        jp = safe_num(row[col_map["price_jpy"]].value)
        if jp is not None:
            fields[F["price_jpy"]] = jp

        cf = safe_float(row[col_map.get("commission_foreign", 99)].value if col_map.get("commission_foreign") is not None else None)
        if cf is not None:
            fields[F["commission_foreign"]] = cf

        cj = safe_num(row[col_map.get("commission_jpy", 99)].value if col_map.get("commission_jpy") is not None else None)
        if cj is not None:
            fields[F["commission_jpy"]] = cj

        tax = safe_num(row[col_map["tax"]].value)
        if tax is not None:
            fields[F["tax"]] = tax

        ship_p = safe_num(row[col_map["shipping_purchase"]].value)
        if ship_p is not None:
            fields[F["shipping_purchase"]] = ship_p

        other_p = safe_num(row[col_map.get("other_purchase", 99)].value if col_map.get("other_purchase") is not None else None)
        if other_p is not None:
            fields[F["other_purchase"]] = other_p

        total_p = safe_num(row[col_map["total_purchase"]].value)
        if total_p is not None:
            fields[F["total_purchase"]] = total_p

        # 販売データ
        sale_date = safe_str(row[col_map["sale_date"]].value)
        if sale_date and sale_date not in ("#", "#VALUE!"):
            fields[F["sale_date"]] = sale_date

        rev = safe_num(row[col_map["revenue"]].value)
        if rev is not None:
            fields[F["revenue"]] = rev

        sc = safe_num(row[col_map["sale_commission"]].value)
        if sc is not None:
            fields[F["sale_commission"]] = sc

        ship_s = safe_num(row[col_map["shipping_sale"]].value)
        if ship_s is not None:
            fields[F["shipping_sale"]] = ship_s

        other_s = safe_num(row[col_map.get("other_sale", 99)].value if col_map.get("other_sale") is not None else None)
        if other_s is not None:
            fields[F["other_sale"]] = other_s

        total_s = safe_num(row[col_map["total_sale"]].value)
        if total_s is not None:
            fields[F["total_sale"]] = total_s

        profit = safe_num(row[col_map["profit"]].value)
        if profit is not None:
            fields[F["profit"]] = profit

        # タイトル・説明文
        title = safe_str(row[col_map["title"]].value)
        if title and title not in ("OK",):
            fields[F["title"]] = title[:200]

        desc = safe_str(row[col_map["description"]].value)
        if desc and desc not in ("OK",):
            fields[F["description"]] = desc[:200]

        # 価格設定
        min_p = safe_num(row[col_map["min_price"]].value)
        if min_p is not None:
            fields[F["min_price"]] = min_p

        min_pr = safe_num(row[col_map["min_profit"]].value)
        if min_pr is not None:
            fields[F["min_profit"]] = min_pr

        buy_p = safe_num(row[col_map["buynow_price"]].value)
        if buy_p is not None:
            fields[F["buynow_price"]] = buy_p

        buy_pr = safe_num(row[col_map["buynow_profit"]].value)
        if buy_pr is not None:
            fields[F["buynow_profit"]] = buy_pr

        # ステータス
        if is_sold:
            fields[F["status"]] = "売却済"
        else:
            status_val = row[col_map.get("status_col", 99)].value if col_map.get("status_col") is not None else None
            fields[F["status"]] = parse_status(status_val)

        # eBay URL
        ebay = safe_str(row[col_map.get("ebay_url", 99)].value if col_map.get("ebay_url") is not None else "")
        if ebay and ebay.startswith("http"):
            fields[F["ebay_url"]] = ebay

        # メモ
        memo = safe_str(row[col_map.get("memo", 99)].value if col_map.get("memo") is not None else "")
        if memo:
            fields[F["memo"]] = memo[:200]

        # 写真
        photo = safe_str(row[col_map.get("photo", 99)].value if col_map.get("photo") is not None else "")
        if photo:
            fields[F["photo"]] = photo[:50]

        records.append({"fields": fields})
        count += 1

    # 10件ずつバッチ投入
    success = 0
    for i in range(0, len(records), 10):
        batch = records[i:i+10]
        r = api_post(batch)
        if r:
            success += len(r.get("records", []))
        time.sleep(0.3)

    print(f"  {sheet_name}: {success}/{count}件移行完了")
    return success


def main():
    print("Excel → Airtable 移行開始")
    print(f"ファイル: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 全体管理シート (ヘッダー行2-3, データ行4〜)
    # B=NO, C=写真/URL, D=素材, E=鑑定番号, F=アイテム名
    # G=仕入日, H=仕入先, I=通貨, J=為替, K=外貨, L=JPY, M=落札手数料外貨, N=落札手数料JPY
    # O=消費税, P=送料, Q=その他, R=合計
    # S=販売日, T=売上, U=販売手数料, V=送料, W=その他, X=合計
    # Y=利益, Z=タイトル, AA=説明文, AB=最低販売値, AC=最低販売時想定利益
    # AD=即決価格, AE=即決価格時想定利益, AF=写真撮影, AG=出品手配, AH=メモ
    col_map_main = {
        "no": 1,          # B
        "ebay_url": 2,    # C
        "material": 3,    # D
        "cert_no": 4,     # E
        "item": 5,        # F
        "source": 7,      # H
        "currency": 8,    # I
        "exchange_rate": 9, # J
        "price_foreign": 10, # K
        "price_jpy": 11,  # L
        "commission_foreign": 12, # M
        "commission_jpy": 13,     # N
        "tax": 14,        # O
        "shipping_purchase": 15, # P
        "other_purchase": 16,    # Q
        "total_purchase": 17,    # R
        "sale_date": 18,  # S
        "revenue": 19,    # T
        "sale_commission": 20, # U
        "shipping_sale": 21,   # V
        "other_sale": 22,      # W
        "total_sale": 23,      # X
        "profit": 24,     # Y
        "title": 25,      # Z
        "description": 26, # AA
        "min_price": 27,  # AB
        "min_profit": 28, # AC
        "buynow_price": 29, # AD
        "buynow_profit": 30, # AE
        "photo": 31,      # AF
        "status_col": 32, # AG
        "memo": 33,       # AH
    }

    ws1 = wb["全体管理"]
    s1 = migrate_sheet(ws1, "全体管理", 4, col_map_main, is_sold=False)

    # 売却済みシート (ヘッダー行2-3, データ行5〜)
    # B=NO, C=写真/URL, D=素材, E=アイテム名
    # F=仕入日, G=仕入先, H=通貨, I=為替, J=外貨, K=JPY, L=落札手数料外貨, M=落札手数料JPY
    # N=消費税, O=送料, P=その他, Q=合計
    # R=販売日, S=売上, T=販売手数料, U=送料, V=その他, W=合計
    # X=利益, Y=タイトル, Z=説明文, AA=最低販売値, AB=最低販売時想定利益
    # AC=即決価格, AD=即決価格時想定利益, AE=写真撮影, AF=出品手配, AG=メモ
    col_map_sold = {
        "no": 1,          # B
        "ebay_url": 2,    # C
        "material": 3,    # D
        "cert_no": None,
        "item": 4,        # E
        "source": 6,      # G
        "currency": 7,    # H
        "exchange_rate": 8, # I
        "price_foreign": 9, # J
        "price_jpy": 10,  # K
        "commission_foreign": 11, # L
        "commission_jpy": 12,     # M
        "tax": 13,        # N
        "shipping_purchase": 14, # O
        "other_purchase": 15,    # P
        "total_purchase": 16,    # Q
        "sale_date": 17,  # R
        "revenue": 18,    # S
        "sale_commission": 19, # T
        "shipping_sale": 20,   # U
        "other_sale": 21,      # V
        "total_sale": 22,      # W
        "profit": 23,     # X
        "title": 24,      # Y
        "description": 25, # Z
        "min_price": 26,  # AA
        "min_profit": 27, # AB
        "buynow_price": 28, # AC
        "buynow_profit": 29, # AD
        "photo": 30,      # AE
        "status_col": 31, # AF
        "memo": 32,       # AG
    }

    ws2 = wb["売却済み"]
    s2 = migrate_sheet(ws2, "売却済み", 5, col_map_sold, is_sold=True)

    print(f"\n=== 移行完了: {s1 + s2}件 ===")


if __name__ == "__main__":
    main()
